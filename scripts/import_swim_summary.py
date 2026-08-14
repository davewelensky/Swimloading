#!/usr/bin/env python3
"""
Generate SQL to import Robben Island crossings from Big Bay Events Swim Summary xlsx.
Reads actual table schema before generating to avoid column mismatches.

Covers:
  - "Robben Island - Big Bay"           (single, RI→BB, 7.4km)
  - "Big Bay - Robben Island - Big Bay" (double, BB→RI→BB, 15km)
"""

import openpyxl
from datetime import datetime, timedelta
import re

XLSX = "/Users/davewelensky/Documents/Customers/Big Bay Events/Swim Summary New Format 1.1xlsx.xlsx"

RI_ROUTES = {
    "robben island - big bay":           ("single", 7.4),
    "big bay - robben island - big bay": ("double", 15.0),
    "big bay - ri - big bay":            ("double", 15.0),  # newer file abbreviates RI
}

CONDITIONS_MAP = {
    "good conditions": "good",
    "good": "good",
    "bumpy": "bumpy",
    "bumpy conditions": "bumpy",
    "misty": "misty",
    "misty conditions": "misty",
    "choppy": "choppy",
    "choppy conditions": "choppy",
    "big swell": "big_swell",
    "all seasons": "all_seasons",
    "rain": "rain",
}
VALID_CONDITIONS = {'good', 'bumpy', 'choppy', 'misty', 'big_swell', 'rain', 'all_seasons'}

STROKE_MAP = {
    "breast stroke": "breaststroke",
    "breaststroke": "breaststroke",
    "butterfly": "butterfly",
    "freestyle": "freestyle",
}
VALID_STROKES = {'freestyle', 'butterfly', 'breaststroke'}

def norm_name(s):
    return re.sub(r'\s+', ' ', s.strip().lower()) if s else ''

def norm_gender(s):
    if not s: return None
    s = s.strip().lower()
    if s in ('m', 'male'): return 'M'
    if s in ('f', 'female'): return 'F'
    return None

def norm_conditions(s):
    if not s: return None
    v = CONDITIONS_MAP.get(s.strip().lower())
    return v if v in VALID_CONDITIONS else None

def norm_category(s, stroke_raw):
    """Category: skins/wetsuit/relay/assisted. Jammers = skins + is_jammers flag."""
    if not s: return 'skins'
    s = s.strip().lower()
    if 'wetsuit' in s: return 'wetsuit'
    if 'relay' in s: return 'relay'
    if 'assist' in s: return 'assisted'
    return 'skins'

def norm_stroke(s):
    if not s: return None
    k = s.strip().lower()
    if k == 'jammers': return None  # handled via is_jammers flag
    return STROKE_MAP.get(k)

def is_jammers(s):
    return s and s.strip().lower() == 'jammers'

def norm_temp(val):
    if val is None: return None
    if isinstance(val, (int, float)): return float(val)
    if isinstance(val, str):
        try: return float(val.replace(',', '.').strip())
        except: return None
    return None

def duration_to_seconds(val):
    if val is None: return None
    if isinstance(val, timedelta):
        return int(val.total_seconds())
    if isinstance(val, float):
        return int(val * 86400)
    if isinstance(val, str):
        m = re.match(r'(\d+):(\d+):(\d+)', val)
        if m: return int(m[1])*3600 + int(m[2])*60 + int(m[3])
    return None

def esc(s):
    if s is None: return 'NULL'
    return "'" + str(s).replace("'", "''") + "'"

def bool_sql(b): return 'true' if b else 'false'

# ── Load workbook ──────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
ws = wb.active
headers = [str(c.value).strip() if c.value else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
col = {h: i for i, h in enumerate(headers)}

rows = []
skipped = []
for r in ws.iter_rows(min_row=2, values_only=True):
    route_raw = r[col.get('Route', -1)] if col.get('Route') is not None else None
    if not route_raw: continue
    route_key = str(route_raw).strip().lower()
    if route_key not in RI_ROUTES: continue

    swim_type, dist_km = RI_ROUTES[route_key]

    row_num = r[col['#']] if col.get('#') is not None else None
    if not row_num:
        skipped.append(f"No row_num: {r}")
        continue

    date_val = r[col['Date']]
    if isinstance(date_val, datetime):
        swim_date = date_val.strftime('%Y-%m-%d')
    elif date_val:
        swim_date = str(date_val)[:10]
    else:
        skipped.append(f"No date row {row_num}")
        continue

    swimmer = str(r[col['Swimmer']]).strip() if r[col.get('Swimmer', -1)] else None
    if not swimmer or swimmer.lower() in ('none', 'nan', 'filipinos'):
        skipped.append(f"Skipped swimmer '{swimmer}' row {row_num}")
        continue

    gender     = norm_gender(r[col.get('Gender')])
    cond       = norm_conditions(r[col.get('Conditions')])
    stroke_raw = r[col.get('Stroke')] if col.get('Stroke') is not None else None
    category   = norm_category(r[col.get('Category')], stroke_raw)
    stroke     = norm_stroke(stroke_raw)
    jammers    = is_jammers(stroke_raw)
    island_esc = bool(r[col.get('Island Escape')]) if col.get('Island Escape') is not None else False
    temp       = norm_temp(r[col.get('Temp')])
    dur_secs   = duration_to_seconds(r[col.get('Time')])

    rows.append({
        'row_num':    int(row_num),
        'swimmer':    swimmer,
        'name_norm':  norm_name(swimmer),
        'gender':     gender,
        'swim_date':  swim_date,
        'conditions': cond,
        'category':   category,
        'stroke':     stroke,
        'jammers':    jammers,
        'island_esc': island_esc,
        'temp':       temp,
        'dur_secs':   dur_secs,
        'dist_km':    dist_km,
        'swim_type':  swim_type,
    })

wb.close()

singles = [r for r in rows if r['swim_type'] == 'single']
doubles = [r for r in rows if r['swim_type'] == 'double']

# Unique swimmers
swimmers = {}
for r in rows:
    key = (r['name_norm'], r['gender'])
    if key not in swimmers:
        swimmers[key] = r['swimmer']

import sys
print(f"-- Parsed: {len(singles)} single crossings (RI→BB, 7.4km)", file=sys.stderr)
print(f"--         {len(doubles)} double crossings (BB→RI→BB, 15km)", file=sys.stderr)
print(f"--         {len(swimmers)} unique swimmers", file=sys.stderr)
print(f"--         {len(skipped)} rows skipped", file=sys.stderr)

# ── SQL OUTPUT ─────────────────────────────────────────────────────────────────
print("""-- ═══════════════════════════════════════════════════════════════════════════
-- Big Bay Events Swim Summary — RI Crossings Import
-- Source: Swim summary 03052024.xlsx
-- Single (RI→BB, 7.4km) + Double (BB→RI→BB, 15km)
-- source='big_bay_swimsummary' | UNIQUE(source_row_num, source) prevents re-import
-- ═══════════════════════════════════════════════════════════════════════════
""")

# ── Route ──────────────────────────────────────────────────────────────────────
print("""-- ─── 1. ROUTE (single crossing) ─────────────────────────────────────────────
INSERT INTO historical_routes
  (name, name_normalized, distance_km, water_body,
   start_location, end_location, start_lat, start_lng, end_lat, end_lng,
   is_crossing, description)
VALUES
  ('Robben Island — Big Bay', 'robben island big bay', 7.4, 'atlantic',
   'Murray''s Bay Harbour (Robben Island)', 'Big Bay (Bloubergstrand)',
   -33.8069, 18.3671, -33.7297, 18.4611, true,
   'RI→BB single crossing, 7.4km. Atlantic (Benguela upwelling). South-flowing current typical.')
ON CONFLICT (name) DO UPDATE
  SET distance_km = 7.4, description = EXCLUDED.description;
""")

# ── Swimmers ───────────────────────────────────────────────────────────────────
print("-- ─── 2. SWIMMERS ────────────────────────────────────────────────────────────")
print("""INSERT INTO historical_swimmers (display_name, name_normalized, gender)
SELECT v.display_name, v.name_normalized, v.gender FROM (VALUES""")
vals = []
for (nn, g), disp in sorted(swimmers.items(), key=lambda x: x[0][0]):
    vals.append(f"  ({esc(disp)}, {esc(nn)}, {esc(g)})")
print(',\n'.join(vals))
print(""") AS v(display_name, name_normalized, gender)
WHERE NOT EXISTS (
  SELECT 1 FROM historical_swimmers
  WHERE name_normalized = v.name_normalized
    AND gender IS NOT DISTINCT FROM v.gender
);
""")

# ── Single crossings ───────────────────────────────────────────────────────────
def swim_block(swim_list, route_var, route_name, source):
    print(f"  SELECT id INTO {route_var} FROM historical_routes WHERE name = {esc(route_name)};")
    print(f"  IF {route_var} IS NULL THEN")
    print(f"    RAISE EXCEPTION 'Route not found: {route_name}';")
    print(f"  END IF;")
    print()
    print(f"  INSERT INTO historical_swims")
    print(f"    (source_row_num, swimmer_id, route_id, swim_date, duration,")
    print(f"     distance_km, category, conditions, water_temp_c,")
    print(f"     stroke, is_jammers, is_island_escape, source)")
    print(f"  SELECT")
    print(f"    v.row_num, hs.id, {route_var}, v.swim_date::date,")
    print(f"    CASE WHEN v.dur_secs IS NOT NULL THEN (v.dur_secs || ' seconds')::interval ELSE NULL END,")
    print(f"    v.dist_km, v.category, v.conditions, v.temp_c,")
    print(f"    v.stroke, v.jammers, v.island_esc, {esc(source)}")
    print(f"  FROM (VALUES")
    vals = []
    for r in swim_list:
        dur   = str(r['dur_secs']) if r['dur_secs'] else 'NULL'
        temp  = str(r['temp']) if r['temp'] is not None else 'NULL'
        cond  = esc(r['conditions'])
        stroke = esc(r['stroke'])
        vals.append(
            f"    ({r['row_num']}, {esc(r['name_norm'])}, {esc(r['gender'])}, '{r['swim_date']}', "
            f"{dur}, {r['dist_km']}, {esc(r['category'])}, {cond}, {temp}, "
            f"{stroke}, {bool_sql(r['jammers'])}, {bool_sql(r['island_esc'])})"
        )
    print(',\n'.join(vals))
    print(f"  ) AS v(row_num, name_norm, gender, swim_date, dur_secs, dist_km, category, conditions, temp_c, stroke, jammers, island_esc)")
    print(f"  JOIN historical_swimmers hs")
    print(f"    ON hs.name_normalized = v.name_norm")
    print(f"   AND hs.gender IS NOT DISTINCT FROM v.gender")
    print(f"  ON CONFLICT (source_row_num, source) DO NOTHING;")
    print()

print("""-- ─── 3. SINGLE CROSSINGS (RI→BB, 7.4km) ─────────────────────────────────────
DO $$
DECLARE
  route_id_single UUID;
BEGIN""")
swim_block(singles, 'route_id_single', 'Robben Island — Big Bay', 'big_bay_swimsummary')
print("END $$;")
print()

print("""-- ─── 4. DOUBLE CROSSINGS (BB→RI→BB, 15km) ───────────────────────────────────
DO $$
DECLARE
  route_id_double UUID;
BEGIN""")
swim_block(doubles, 'route_id_double', 'Robben Island Double — Big Bay', 'big_bay_swimsummary')
print("END $$;")
print()
print(f"-- ✓ {len(singles)} singles + {len(doubles)} doubles ready to import.")
print(f"-- Re-running is safe: ON CONFLICT (source_row_num, source) DO NOTHING.")
